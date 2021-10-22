import os.path
import pandas as pd

from openpyxl import load_workbook
import xlrd

from datetime import datetime
import dateutil.parser as dparser
from abc import abstractmethod

from shapely.geometry import Point

# Define o tipo de dado contendo uma seleção de células do excel
class range_cells:
    def __init__(self, cols, row_init, n_rows):
        self.cols = cols
        self.row_init = row_init
        self.n_rows = n_rows

# Responsável por realizar operações em arquivos excel
class Excel:
    def __init__(self, file):
        try: 
            os.path.isfile(file)
            self.__file = file
            self.__wb = load_workbook(file).active
        except:
            print('Não foi possível iniciar o arquivo ' + file)

    def get_workbook(self):
        return load_workbook(self.__file).active

    # Retorna um DataFrame a partir do path do excel e da seleção de células que contém a tabela
    def get_table_as_dataframe(self, range_cells, sheet_name):
        cols = range_cells.cols
        row_init = range_cells.row_init
        n_rows = range_cells.n_rows
        data = pd.read_excel(self.__file, sheet_name=sheet_name, skiprows=row_init, nrows=n_rows,usecols=cols)
        data.columns = data.columns.map(str.strip)
        return data

    def get_cell_value(self, col: str, row: str):
        return self.__wb[col + row].value

    @abstractmethod
    def get_data(self):
        pass

    @property
    def file(self):
        return self.__file
    
    @property
    def wb(self):
        return self.__wb

class Excel_Ipdo(Excel):
    def __init__(self, file):
        super(Excel_Ipdo, self).__init__(file)
        self.__wb = super(Excel_Ipdo, self).get_workbook()

    def get_data(self):
        return datetime.strptime(self.__wb['T6'].value, "%d/%m/%Y")

    def get_IPDO_dataframe(self):
        row_init, nrows = 59, 6
        cols = 'K,M,O,P,Q,R'
        # Atribuição em 'earm, earm %, ena medio'
        tabela = super().get_table_as_dataframe(range_cells(cols, row_init, nrows), 'IPDO')
        tabela = tabela.drop(tabela.index[0], axis=0).reset_index(drop=True)
        # Atribuição em 'carga prevista'
        tabela['Carga prevista'] = pd.Series([
            self.get_cell_value('M','24'), # Carga prev. Norte ATUAL
            self.get_cell_value('M','32'), # Carga prev. Nordeste ATUAL
            self.get_cell_value('M','48'), # Carga prev. Sul ATUAL
            self.get_cell_value('M','40'), # Carga prev. Sudeste ATUAL 
            self.get_cell_value('M','16')  # Carga prev. SIN ATUAL
            ])

        # Atribuição em 'carga realizada'
        tabela['Carga realizada'] = pd.Series([
            self.get_cell_value('O','24'), # Carga real. Norte ATUAL
            self.get_cell_value('O','32'), # Carga real. Nordeste ATUAL
            self.get_cell_value('O','48'), # Carga real. Sul ATUAL
            self.get_cell_value('O','40'), # Carga real. Sudeste ATUAL 
            self.get_cell_value('O','16')  # Carga real. SIN
            ])
        
        # Atribuição em 'geração eólica'
        tabela['geracao_eolica'] = pd.Series([
            self.get_cell_value('O','21'), # Geração eólica Norte ATUAL
            self.get_cell_value('O','29'), # Geração eólica Nordeste ATUAL
            self.get_cell_value('O','45'), # Geração eólica Sul ATUAL
            self.get_cell_value('O','37'), # Geração eólica Sudeste ATUAL
            self.get_cell_value('O','12')  # Geração eólica SIN ATUAL
            ])

        # Atribuição em 'geração solar'
        tabela['geracao_solar'] = pd.Series([
            self.get_cell_value('O','22'), # Geração solar Norte ATUAL
            self.get_cell_value('O','30'), # Geração solar Nordeste ATUAL
            self.get_cell_value('O','46'), # Geração solar Sul ATUAL
            self.get_cell_value('O','38'), # Geração solar Sudeste ATUAL
            self.get_cell_value('O','13')  # Geração solar SIN ATUAL
            ])

        tabela['Data'] = self.get_data()
        tabela.rename(columns={tabela.columns[0]: 'regiao',
                                tabela.columns[1]: 'ena_mwmed',
                                tabela.columns[2]: 'ena_bruta_percentual',
                                tabela.columns[3]: 'ena_armaz_percentual',
                                tabela.columns[4]: 'earm',
                                tabela.columns[5]: 'earm_percentual',
                                tabela.columns[6]: 'carga_prevista',
                                tabela.columns[7]: 'carga_realizada',
                                tabela.columns[8]: 'geracao_eolica',
                                tabela.columns[9]: 'geracao_solar',
                                tabela.columns[10]: 'data',
                                }, inplace=True)
        tabela.loc[tabela.regiao == 'SIN', 'ena_mwmed'] = sum(tabela.ena_mwmed[:-1])                
        tabela[['earm_percentual', 'ena_armaz_percentual', 'ena_bruta_percentual']] *= 100
        cols = ['regiao', 'data', 'ena_mwmed', 'ena_bruta_percentual', 'ena_armaz_percentual',
        'earm', 'earm_percentual', 'carga_prevista', 'carga_realizada', 'geracao_eolica', 'geracao_solar']
        tabela = tabela[cols]
        tabela = tabela.round(2)
        return tabela

class Excel_Acomph():
    def __init__(self, file):
        try:
            os.path.isfile(file)
            self.__file = file
            self.__wb = xlrd.open_workbook(file)
        except:
            print('Não foi possível iniciar o arquivo ' + file)

    def get_dataframe_by_sheet(self, sheet: int):
        sh = self.wb.sheet_by_index(sheet)
        cod_posto_column = 8

        df = pd.DataFrame()

        # Iteração para pegar todos os postos da planilha 'sheet'
        while cod_posto_column < sh.ncols:
            df_posto_30_dias = pd.DataFrame(columns=['cod_posto', 'data', 'vazao_natural', 'vazao_incremental'])
            cod_posto = int(sh[0,cod_posto_column].value)
            for i in range(5,35):
                data = xlrd.xldate.xldate_as_datetime(sh[i,0].value, 0)
                vazao_natural = sh[i,cod_posto_column].value
                vazao_incremental = sh[i,cod_posto_column-1].value
                series = pd.Series(data={'cod_posto': cod_posto,
                               'data': data,
                               'vazao_natural': vazao_natural,
                               'vazao_incremental': vazao_incremental,
                               })
                df_posto_30_dias = df_posto_30_dias.append(series, ignore_index=True)
            df = df.append(df_posto_30_dias, ignore_index=True).reset_index(drop=True)
            cod_posto_column += 8

        df = df.reset_index(drop=True).round(2)
        return df

    @property
    def file(self):
        return self.__file
    
    @property
    def wb(self):
        return self.__wb

class Excel_RDH(Excel):
    def __init__(self, file):
        try:
            super(Excel_RDH, self).__init__(file)
            self.__wb = load_workbook(file).worksheets[0]
        except:
            print('Não foi possível iniciar o arquivo ' + file)

    def get_rdh_dataframe(self):
        df = pd.DataFrame()
        data = dparser.parse(self.wb['U2'].value, fuzzy=True, dayfirst=True).date()
        for i in range(8,173):
            cod_posto = self.wb['E' + str(i)].value
            nivel = self.wb['O' + str(i)].value
            volume_perc = self.wb['P' + str(i)].value
            df = df.append(pd.Series([cod_posto,data,nivel,volume_perc]), ignore_index=True)

        df = df.rename(columns={0: 'cod_posto', 1: 'data', 2: 'nivel', 3: 'volume_perc'})
        df.drop(df[df.cod_posto == 'ND'].index, inplace=True)
        df.cod_posto = pd.to_numeric(df.cod_posto,errors='coerce').convert_dtypes()
        df.data = pd.to_datetime(df.data)
        df.volume_perc = pd.to_numeric(df.volume_perc,errors='coerce').convert_dtypes()

        # Posto E.SOUZA desativado
        df.drop(df[df.cod_posto == 164].index, inplace=True)

        # Posto P.A.123 desativado
        df.drop(df[df.cod_posto == 174].index, inplace=True)
        return df

    @property
    def file(self):
        return self.__file
    
    @property
    def wb(self):
        return self.__wb

class Precipitacao():
    def __init__(self, file):
        self.__file = file

    def get_data_arquivo(self, type):
        if type == 'gefs':
            return datetime.strptime(os.path.basename(self.file)[6:12], '%d%m%y')
        if type == 'eta' or type == 'ecmwf':
            return datetime.strptime(os.path.basename(self.file)[7:13], '%d%m%y')
        raise('ERRO. Nome de arquivo inválido!') 

    def get_data_precipitacao(self, type):
        if type == 'gefs':
            return datetime.strptime(os.path.basename(self.file)[13:19], '%d%m%y')
        if type == 'eta' or type == 'ecmwf':
            return datetime.strptime(os.path.basename(self.file)[14:20], '%d%m%y')
        raise('ERRO. Nome de arquivo inválido!') 

    def get_dataframe(self):
        return pd.read_csv(self.file, names=['long', 'lat', 'precipitacao'], sep="\s+")
    

    
    @property
    def file(self):
        return self.__file